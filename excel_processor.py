import openpyxl
from openpyxl.chart import BarChart, Reference
import matplotlib.pyplot as plt


# 1. Загрузка файла
def load_data(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append({"name": row[0], "price": row[1], "quantity": row[2]})
    return data, wb, sheet


# 2. Сортировка по цене (от высокой к низкой)
def sort_data(data):
    return sorted(data, key=lambda x: x["price"], reverse=True)


# 3. Сохранение в новый файл + график
def save_sorted_data(data, wb, sheet):
    # Очистка старых данных
    for row in sheet.iter_rows(min_row=2, max_col=3, max_row=len(data) + 1):
        for cell in row:
            cell.value = None

    # Запись новых данных
    for idx, item in enumerate(data, start=2):
        sheet.cell(row=idx, column=1, value=item["name"])
        sheet.cell(row=idx, column=2, value=item["price"])
        sheet.cell(row=idx, column=3, value=item["quantity"])

    # Добавляем график в Excel
    chart = BarChart()
    chart.title = "Топ товаров по цене"
    chart.x_axis.title = "Товары"
    chart.y_axis.title = "Цена"

    data_ref = Reference(sheet, min_col=2, min_row=1, max_row=len(data) + 1, max_col=2)
    categories_ref = Reference(sheet, min_col=1, min_row=2, max_row=len(data) + 1)

    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(categories_ref)
    sheet.add_chart(chart, "E2")

    wb.save("sorted_data.xlsx")


# 4. График в matplotlib (альтернатива)
def plot_matplotlib(data):
    names = [item["name"] for item in data]
    prices = [item["price"] for item in data]

    plt.bar(names, prices, color='skyblue')
    plt.title("Топ товаров по цене")
    plt.xlabel("Товары")
    plt.ylabel("Цена (руб)")
    plt.xticks(rotation=45)
    plt.tight_layout()
    plt.savefig("plot.png")  # Сохраняем картинку
    plt.show()


# Запуск
if __name__ == "__main__":
    data, wb, sheet = load_data("data.xlsx")
    sorted_data = sort_data(data)
    save_sorted_data(sorted_data, wb, sheet)
    plot_matplotlib(sorted_data)